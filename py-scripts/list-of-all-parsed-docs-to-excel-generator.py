import json
import pandas as pd
import os
from pathlib import Path

# ============================================
# CONFIGURATION - ALL JSON FILES IN ONE PLACE
# ============================================
JSON_FILES_CONFIG = {
    'gcc_agreements': [
        "0-gcc-customs-agreement.json",
        "1-gcc-vat-agreement.json",
        "2-gcc-excise-agreement.json"
    ],
    'uae_laws': [
        "3-uae-cit-47-country-law-articles-decisions.json",
        "6-uae-vat-country-law-articles-decisions.json",
        "8-uae-tp-country-law-articles.json",
        "10-uae-excise-country-law-articles.json",
        "12-uae-fatca-law.json",
        "temp-uae-excise-country-law-articles.json"
    ],
    'ksa_laws': [
        "13-ksa-incometax-country-law-articles-guides.json",
        "14-ksa-vat-country-law-articles-decisions.json",
        "16-ksa-excise-country-law-articles.json",
        "17-ksa-zakat-country-law-articles.json",
        "34-ksa-customs-country-law-guides.json",
        "35-ksa-fatca-law.json"
    ],
    'kuwait_laws': [
        "19-kwt-bptl-country-law-articles.json",
        "20-kwt-ktl-country-law-articles.json",
        "21-kwt-dl-157-country-law-articles.json"
    ],
    'qatar_laws': [
        "22-qatar-incometax-country-law-articles-decisions.json",
        "24-qatar-tp-country-law-articles-decisions-circulars.json",
        "25-qatar-excisetax-country-law-articles-decisions.json",
        "26-qatar-fatca-law.json"
    ],
    'bahrain_laws': [
        "27-bahrain-incometax-country-law-articles.json"
    ],
    'oman_laws': [
        "28-oman-incometax-country-law-articles.json",
        "29-oman-vat-country-law-articles-decisions.json",
        "31-oman-excise-country-law-guides.json",
        "32-oman-cbcr-country-law-guides.json",
        "33-oman-fatca-law.json"
    ],
    'uae_guidelines': [
        "4-uae-cit-guidelines-guide.json",
        "4-uae-cit-guidelines-pc.json",
        "7-uae-vat-guidelines-guide-pc.json",
        "9-uae-tp-guidelines-guide-pc.json",
        "11-uae-excise-guidelines-guide.json",
        "11-uae-excise-guidelines-pc.json"
    ],
    'ksa_guidelines': [
        "15-ksa-vat-guidelines-guide.json",
        "18-ksa-zakat-guidelines-guide.json",
        "36-ksa-incometax-guides.json"
    ],
    'qatar_guidelines': [
        "23-qatar-incometax-circulars.json"
    ],
    'oman_guidelines': [
        "30-oman-vat-guidelines-guide.json"
    ],
    'dtaa_agreements': [
        "dtaa-uae-1.json",
        "dtaa-uae-2.json",
        "dtaa-ksa.json",
        "dtaa-kuwait.json",
        "dtaa-qatar.json",
        "dtaa-oman.json",
        "dtaa-bahrain.json"
    ],
    'blogs': [
        "blogs.json"
    ]
}

# Country code mapping
COUNTRY_CODES = {
    'uae': 'UAE',
    'ksa': 'KSA',
    'kwt': 'Kuwait',
    'kuwait': 'Kuwait',
    'omn': 'Oman',
    'oman': 'Oman',
    'bhr': 'Bahrain',
    'bahrain': 'Bahrain',
    'qat': 'Qatar',
    'qatar': 'Qatar'
}

# Document type processing order
DOC_TYPE_ORDER = ['Article', 'Decision', 'Guideline', 'Circular', 'DTAA']

def read_json_file(file_path):
    """Read and parse JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"  ✗ Error reading {file_path}: {e}")
        return None

def extract_article_number(doc_name):
    """Extract numeric article number for proper sorting"""
    import re
    match = re.search(r'\d+', str(doc_name))
    if match:
        return int(match.group(0))
    return float('inf')

def extract_dtaa_title(dtaa_obj):
    """Extract DTAA title from a single DTAA object"""
    country1 = dtaa_obj.get('country1Slug', '').upper()
    country2 = dtaa_obj.get('country2Name', '')
    return f"{country1} - {country2}" if country1 and country2 else "Unknown DTAA"

def get_country_from_filename(filename):
    """Determine country from filename"""
    filename_lower = filename.lower()
    for code, name in COUNTRY_CODES.items():
        if code in filename_lower:
            return name
    return None

def extract_sheet_name_from_file(file_name):
    """Extract sheet name from filename"""
    name = file_name.replace('.json', '').lower()
    if name[0].isdigit() and '-' in name:
        name = name[name.index('-')+1:]
    return name

def process_gcc_agreements(data_dir, file_list, gcc_sheets):
    """Process GCC agreement files with detailed articles and decisions"""
    print("\n📋 Processing GCC Agreements...")
    
    for file_name in file_list:
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            print(f"  ⚠ File not found: {file_name}")
            continue
        
        data = read_json_file(file_path)
        if not data:
            continue
        
        print(f"  Processing: {file_name}")
        
        sheet_name = extract_sheet_name_from_file(file_name)
        sheet_data = []
        
        title = ''
        if isinstance(data, dict) and 'laws' in data:
            laws = data['laws']
            if isinstance(laws, list) and len(laws) > 0:
                title = laws[0].get('lawFullName', '')
                law = laws[0]
                
                article_count = 0
                decision_count = 0
                
                # Process Articles
                articles = law.get('articles', [])
                for article in articles:
                    if isinstance(article, dict):
                        article_number = article.get('number', '')
                        article_title = article.get('title', '') or 'Unknown'
                        
                        sheet_data.append({
                            'Item Type': 'Article',
                            'Item Number': article_number,
                            'Item Title': article_title
                        })
                        article_count += 1
                
                # Process Decisions
                decisions = law.get('decisions', [])
                for decision in decisions:
                    if isinstance(decision, dict):
                        decision_title = decision.get('title', '') or 'Unknown'
                        decision_name = decision.get('name', '') or 'Unknown'
                        
                        sheet_data.append({
                            'Item Type': 'Decision',
                            'Item Number': decision.get('year', ''),
                            'Item Title': decision_name if decision_name != 'Unknown' else decision_title
                        })
                        decision_count += 1
                
                count_summary = []
                if article_count > 0:
                    count_summary.append(f"{article_count} articles")
                if decision_count > 0:
                    count_summary.append(f"{decision_count} decisions")
                
                print(f"    ✓ Added {title} with {', '.join(count_summary)}")
        
        if sheet_data:
            gcc_sheets[sheet_name] = {
                'law_name': title,
                'articles': sheet_data,
                'file_name': file_name
            }

def process_dtaa_agreements(data_dir, country_data, file_list):
    """Process DTAA agreement files"""
    print("\n🤝 Processing DTAA Agreements...")
    
    for file_name in file_list:
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            print(f"  ⚠ File not found: {file_name}")
            continue
        
        data = read_json_file(file_path)
        if not data:
            continue
        
        print(f"  Processing: {file_name}")
        
        filename_lower = file_name.lower()
        country_code = None
        
        for code in COUNTRY_CODES.keys():
            if code in filename_lower:
                country_code = code
                break
        
        country_name = COUNTRY_CODES.get(country_code, 'Other')
        
        dtaa_list = data if isinstance(data, list) else [data]
        dtaa_count = 0
        
        for dtaa_obj in dtaa_list:
            if not isinstance(dtaa_obj, dict):
                continue
            
            title = extract_dtaa_title(dtaa_obj)
            
            if country_code in COUNTRY_CODES:
                country = COUNTRY_CODES[country_code]
                country_data[country].append({
                    'Document Name': title,
                    'Type': 'DTAA',
                    'File Name': file_name
                })
            
            dtaa_count += 1
        
        print(f"    ✓ Added {dtaa_count} DTAA agreement(s)")

def process_country_laws(data_dir, country_data, file_list, category_name):
    """Process country law files"""
    print(f"\n📜 Processing {category_name}...")
    
    for file_name in file_list:
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            print(f"  ⚠ File not found: {file_name}")
            continue
        
        data = read_json_file(file_path)
        if not data:
            continue
        
        print(f"  Processing: {file_name}")
        
        country = get_country_from_filename(file_name)
        if not country:
            print(f"    ⚠ Could not determine country for {file_name}")
            continue
        
        article_count = 0
        decision_count = 0
        guideline_count = 0
        circular_count = 0
        
        if isinstance(data, dict) and 'laws' in data:
            for law in data['laws']:
                # Process Articles
                if 'articles' in law and isinstance(law['articles'], list):
                    for article in law['articles']:
                        article_title = article.get('title', '') or 'Unknown'
                        country_data[country].append({
                            'Document Name': article_title,
                            'Type': 'Article',
                            'File Name': file_name
                        })
                        article_count += 1
                
                # Process Decisions
                if 'decisions' in law and isinstance(law['decisions'], list):
                    for decision in law['decisions']:
                        decision_name = decision.get('name', '') or 'Unknown'
                        country_data[country].append({
                            'Document Name': decision_name,
                            'Type': 'Decision',
                            'File Name': file_name
                        })
                        decision_count += 1
                
                # Process Guidelines
                if 'guidelines' in law and isinstance(law['guidelines'], list):
                    for guideline in law['guidelines']:
                        guideline_title = guideline.get('title', '') or 'Unknown'
                        country_data[country].append({
                            'Document Name': guideline_title,
                            'Type': 'Guideline',
                            'File Name': file_name
                        })
                        guideline_count += 1
                
                # Process Circulars
                if 'circulars' in law and isinstance(law['circulars'], list):
                    for circular in law['circulars']:
                        circular_title = circular.get('title', '') or 'Unknown'
                        country_data[country].append({
                            'Document Name': circular_title,
                            'Type': 'Circular',
                            'File Name': file_name
                        })
                        circular_count += 1
        
        counts = []
        if article_count > 0:
            counts.append(f"{article_count} Articles")
        if decision_count > 0:
            counts.append(f"{decision_count} Decisions")
        if guideline_count > 0:
            counts.append(f"{guideline_count} Guidelines")
        if circular_count > 0:
            counts.append(f"{circular_count} Circulars")
        
        print(f"    ✓ Added {', '.join(counts) if counts else 'no documents'}")

def process_country_guidelines(data_dir, country_data, file_list, category_name):
    """Process standalone country guideline files"""
    print(f"\n📖 Processing {category_name}...")
    
    for file_name in file_list:
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            print(f"  ⚠ File not found: {file_name}")
            continue
        
        data = read_json_file(file_path)
        if not data:
            continue
        
        print(f"  Processing: {file_name}")
        
        country = get_country_from_filename(file_name)
        if not country:
            print(f"    ⚠ Could not determine country for {file_name}")
            continue
        
        guideline_count = 0
        
        if isinstance(data, list):
            for guide in data:
                guide_title = guide.get('title', '') or 'Unknown'
                country_data[country].append({
                    'Document Name': guide_title,
                    'Type': 'Guideline',
                    'File Name': file_name
                })
                guideline_count += 1
        
        elif isinstance(data, dict):
            if 'guidelines' in data and isinstance(data['guidelines'], list):
                for guideline in data['guidelines']:
                    guide_title = guideline.get('title', '') or 'Unknown'
                    country_data[country].append({
                        'Document Name': guide_title,
                        'Type': 'Guideline',
                        'File Name': file_name
                    })
                    guideline_count += 1
            
            elif 'title' in data:
                guide_title = data.get('title', '') or 'Unknown'
                country_data[country].append({
                    'Document Name': guide_title,
                    'Type': 'Guideline',
                    'File Name': file_name
                })
                guideline_count += 1
        
        print(f"    ✓ Added {guideline_count} Guidelines")

def process_blogs(data_dir, file_list):
    """Process blog files"""
    print(f"\n📝 Processing Blogs...")
    
    blog_data = []
    
    for file_name in file_list:
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            print(f"  ⚠ File not found: {file_name}")
            continue
        
        data = read_json_file(file_path)
        if not data:
            continue
        
        print(f"  Processing: {file_name}")
        
        blog_count = 0
        
        if isinstance(data, list):
            for blog in data:
                blog_title = blog.get('title', '') or 'Unknown'
                blog_data.append({
                    'Blog Title': blog_title,
                    'Category': blog.get('category', '') or 'Uncategorized',
                    'File Name': file_name
                })
                blog_count += 1
        
        elif isinstance(data, dict) and 'blogs' in data:
            for blog in data['blogs']:
                blog_title = blog.get('title', '') or 'Unknown'
                blog_data.append({
                    'Blog Title': blog_title,
                    'Category': blog.get('category', '') or 'Uncategorized',
                    'File Name': file_name
                })
                blog_count += 1
        
        print(f"    ✓ Added {blog_count} Blogs")
    
    return blog_data

def process_documents(data_dir):
    """Process all JSON files based on configuration"""
    
    country_data = {
        'UAE': [],
        'KSA': [],
        'Kuwait': [],
        'Oman': [],
        'Bahrain': [],
        'Qatar': []
    }
    
    gcc_sheets = {}
    
    # Process each category
    process_gcc_agreements(data_dir, JSON_FILES_CONFIG['gcc_agreements'], gcc_sheets)
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['uae_laws'], "UAE Laws")
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['ksa_laws'], "KSA Laws")
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['kuwait_laws'], "Kuwait Laws")
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['qatar_laws'], "Qatar Laws")
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['bahrain_laws'], "Bahrain Laws")
    process_country_laws(data_dir, country_data, JSON_FILES_CONFIG['oman_laws'], "Oman Laws")
    process_country_guidelines(data_dir, country_data, JSON_FILES_CONFIG['uae_guidelines'], "UAE Guidelines")
    process_country_guidelines(data_dir, country_data, JSON_FILES_CONFIG['ksa_guidelines'], "KSA Guidelines")
    process_country_guidelines(data_dir, country_data, JSON_FILES_CONFIG['qatar_guidelines'], "Qatar Guidelines")
    process_country_guidelines(data_dir, country_data, JSON_FILES_CONFIG['oman_guidelines'], "Oman Guidelines")
    process_dtaa_agreements(data_dir, country_data, JSON_FILES_CONFIG['dtaa_agreements'])
    
    blog_data = process_blogs(data_dir, JSON_FILES_CONFIG['blogs'])
    
    return country_data, gcc_sheets, blog_data

def create_excel(country_data, gcc_sheets, blog_data, output_file):
    """Create formatted Excel file with organized country sheets"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ============================================
        # CREATE GCC AGREEMENT SHEETS
        # ============================================
        for sheet_name, sheet_info in gcc_sheets.items():
            if sheet_info['articles']:
                df = pd.DataFrame(sheet_info['articles'])
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                worksheet = writer.sheets[sheet_name]
                apply_formatting(worksheet, df)
                
                worksheet.insert_rows(1)
                worksheet['A1'] = f"Agreement: {sheet_info['law_name']}"
                worksheet['A1'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                worksheet['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                worksheet.column_dimensions['A'].width = 20
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['C'].width = 100
        
        # ============================================
        # CREATE COUNTRY SHEETS
        # ============================================
        for country in ['UAE', 'KSA', 'Kuwait', 'Qatar', 'Bahrain', 'Oman']:
            if not country_data[country]:
                continue
            
            df = pd.DataFrame(country_data[country])
            df = df[['Document Name', 'Type', 'File Name']]
            
            final_data = []
            file_names = sorted(df['File Name'].unique(), key=lambda x: (extract_article_number(x), x))
            
            for file_name in file_names:
                file_docs = df[df['File Name'] == file_name]
                
                final_data.append({
                    'Document Name': f'═══ {file_name} ═══',
                    'Type': '',
                    'File Name': ''
                })
                
                type_order = {t: i for i, t in enumerate(DOC_TYPE_ORDER, 1)}
                file_docs_sorted = file_docs.copy()
                file_docs_sorted['sort_order'] = file_docs_sorted['Type'].map(type_order)
                file_docs_sorted['article_num'] = file_docs_sorted['Document Name'].apply(extract_article_number)
                file_docs_sorted = file_docs_sorted.sort_values(['sort_order', 'article_num', 'Document Name'])
                file_docs_sorted = file_docs_sorted.drop('article_num', axis=1)
                
                current_type = None
                for _, row in file_docs_sorted.iterrows():
                    doc_type = row['Type']
                    
                    if doc_type != current_type:
                        if current_type is not None:
                            final_data.append({
                                'Document Name': '',
                                'Type': '',
                                'File Name': ''
                            })
                        
                        type_count = len(file_docs_sorted[file_docs_sorted['Type'] == doc_type])
                        final_data.append({
                            'Document Name': f'  --- {doc_type}s ({type_count}) ---',
                            'Type': '',
                            'File Name': ''
                        })
                        current_type = doc_type
                    
                    final_data.append({
                        'Document Name': row['Document Name'],
                        'Type': row['Type'],
                        'File Name': row['File Name']
                    })
                
                final_data.append({
                    'Document Name': '',
                    'Type': '',
                    'File Name': ''
                })
            
            final_df = pd.DataFrame(final_data)
            final_df.to_excel(writer, sheet_name=country, index=False)
            
            worksheet = writer.sheets[country]
            apply_formatting(worksheet, final_df)
            
            worksheet.column_dimensions['A'].width = 100
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 35
            
            add_country_summary(worksheet, df, country)
        
        # ============================================
        # CREATE BLOGS SHEET
        # ============================================
        if blog_data:
            df_blogs = pd.DataFrame(blog_data)
            df_blogs.to_excel(writer, sheet_name='Blogs', index=False)
            
            worksheet = writer.sheets['Blogs']
            apply_formatting(worksheet, df_blogs)
            
            worksheet.column_dimensions['A'].width = 100
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 35

def apply_formatting(worksheet, df):
    """Apply consistent formatting to worksheet"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    normal_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', horizontal='left')
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('---'):
                cell.font = section_font
                cell.fill = section_fill
            else:
                cell.font = normal_font

def add_country_summary(worksheet, df, country):
    """Add document count summary to country sheet"""
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    
    col = 5
    row = 1
    
    type_counts = df['Type'].value_counts().to_dict()
    
    worksheet.cell(row=row, column=col, value='SUMMARY')
    worksheet.cell(row=row, column=col).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    worksheet.cell(row=row, column=col).fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for i, doc_type in enumerate(DOC_TYPE_ORDER):
        if doc_type in type_counts:
            row_num = row + i + 1
            worksheet.cell(row=row_num, column=col, value=f'{doc_type}s: {type_counts[doc_type]}')
            worksheet.cell(row=row_num, column=col).font = Font(name='Calibri', size=10)
    
    total = len(df)
    row_num = row + len(DOC_TYPE_ORDER) + 2
    worksheet.cell(row=row_num, column=col, value=f'TOTAL: {total}')
    worksheet.cell(row=row_num, column=col).font = Font(name='Calibri', size=11, bold=True)
    worksheet.cell(row=row_num, column=col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    worksheet.column_dimensions[get_column_letter(col)].width = 20

def main():
    """Main execution function"""
    current_dir = Path(__file__).parent
    data_dir = current_dir / 'data'
    output_file = current_dir / 'documents_inventory.xlsx'
    
    print("=" * 70)
    print("📊 DOCUMENT INVENTORY GENERATOR")
    print("=" * 70)
    print(f"📁 Data directory: {data_dir}")
    print(f"📄 Output file: {output_file}")
    print("=" * 70)
    print(f"\n📝 Configuration: {sum(len(v) for v in JSON_FILES_CONFIG.values())} files configured")
    
    # Process all documents
    country_data, gcc_sheets, blog_data = process_documents(data_dir)
    
    # Print summary
    print("\n" + "=" * 70)
    print("📈 PROCESSING SUMMARY")
    print("=" * 70)
    print("\nCountry Sheets:")
    for country in ['UAE', 'KSA', 'Kuwait', 'Qatar', 'Bahrain', 'Oman']:
        if country_data[country]:
            print(f"  {country:15s}: {len(country_data[country]):4d} documents")
    
    print(f"\nGCC Agreement Sheets: {len(gcc_sheets)}")
    for sheet_name in gcc_sheets.keys():
        print(f"  - {sheet_name}")
    
    print(f"\nBlogs: {len(blog_data)} blog(s)")
    
    # Create Excel file
    print("\n" + "=" * 70)
    print("💾 Creating Excel file...")
    create_excel(country_data, gcc_sheets, blog_data, output_file)
    
    print(f"\n✅ Excel file created successfully!")
    print(f"📍 Location: {output_file}")
    print("=" * 70)

if __name__ == "__main__":
    main()